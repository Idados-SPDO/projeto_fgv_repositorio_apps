import subprocess
import sys

# Ensure openpyxl is installed
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ============================================================
# ESTILOS
# ============================================================
AZUL_ESCURO = "1B2A4A"
AZUL_MEDIO = "2D4A7A"
AZUL_CLARO = "4A90D9"
VERDE = "27AE60"
AMARELO = "F39C12"
VERMELHO = "E74C3C"
CINZA_CLARO = "F2F3F4"
CINZA_MEDIO = "D5D8DC"
BRANCO = "FFFFFF"

header_font = Font(name="Calibri", bold=True, color=BRANCO, size=11)
header_fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=BRANCO, size=10)
subheader_fill = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=AZUL_ESCURO, size=14)
section_font = Font(name="Calibri", bold=True, color=AZUL_ESCURO, size=12)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
zebra_fill = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color=CINZA_MEDIO),
    right=Side(style="thin", color=CINZA_MEDIO),
    top=Side(style="thin", color=CINZA_MEDIO),
    bottom=Side(style="thin", color=CINZA_MEDIO),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = thin_border
            if (row - start_row) % 2 == 1:
                cell.fill = zebra_fill

def auto_width(ws, cols, min_width=12, max_width=40):
    for col in range(1, cols + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len

# ============================================================
# ABA 1: LOG DIÁRIO
# ============================================================
ws1 = wb.active
ws1.title = "Log Diário"
ws1.sheet_properties.tabColor = AZUL_ESCURO

ws1.merge_cells("A1:F1")
ws1["A1"] = "📅 LOG DIÁRIO — Gestão de Dados SPDO"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

headers = ["Data", "O que fiz hoje", "Pendências para amanhã", "Bloqueios", "Decisões tomadas", "Observações"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 6)

ws1.cell(row=4, column=1, value="25/06/2026")
ws1.cell(row=4, column=2, value="Início da transição - organização dos processos e planilhas")
ws1.cell(row=4, column=3, value="Agendar conversas com Brendon e Olivio")
ws1.cell(row=4, column=4, value="Nenhum")
ws1.cell(row=4, column=5, value="Estruturar Planner e planilhas de controle")
ws1.cell(row=4, column=6, value="Primeiro dia assumindo a área")

style_data_rows(ws1, 4, 35, 6)

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 45
ws1.column_dimensions["C"].width = 40
ws1.column_dimensions["D"].width = 30
ws1.column_dimensions["E"].width = 35
ws1.column_dimensions["F"].width = 35

ws1.freeze_panes = "A4"

# ============================================================
# ABA 2: GESTÃO DA EQUIPE
# ============================================================
ws2 = wb.create_sheet("Gestão da Equipe")
ws2.sheet_properties.tabColor = VERDE

ws2.merge_cells("A1:F1")
ws2["A1"] = "👥 GESTÃO DA EQUIPE"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Brendon section
ws2.merge_cells("A3:F3")
ws2["A3"] = "BRENDON — Atividades Operacionais"
ws2["A3"].font = Font(name="Calibri", bold=True, color=BRANCO, size=11)
ws2["A3"].fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
ws2["A3"].alignment = center_align

brendon_headers = ["Data", "Atividade Atribuída", "Status", "Prazo", "Feedback / Observações", "Próximos Passos"]
for i, h in enumerate(brendon_headers, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 6)
style_data_rows(ws2, 5, 14, 6)

# Olivio section
ws2.merge_cells("A16:F16")
ws2["A16"] = "OLIVIO — Apoio ao Flavio (Governança / Banco de Preços)"
ws2["A16"].font = Font(name="Calibri", bold=True, color=BRANCO, size=11)
ws2["A16"].fill = PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid")
ws2["A16"].alignment = center_align

olivio_headers = ["Data", "Atividade", "Status", "Prazo", "Observações", "Ponto de Atenção"]
for i, h in enumerate(olivio_headers, 1):
    ws2.cell(row=17, column=i, value=h)
style_header_row(ws2, 17, 6)
style_data_rows(ws2, 18, 27, 6)

# Reuniões 1:1
ws2.merge_cells("A29:F29")
ws2["A29"] = "REUNIÕES 1:1 / ALINHAMENTOS"
ws2["A29"].font = Font(name="Calibri", bold=True, color=BRANCO, size=11)
ws2["A29"].fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
ws2["A29"].alignment = center_align

reuniao_headers = ["Data", "Com quem", "Pauta", "Decisões", "Ações Combinadas", "Próxima Reunião"]
for i, h in enumerate(reuniao_headers, 1):
    ws2.cell(row=30, column=i, value=h)
style_header_row(ws2, 30, 6)
style_data_rows(ws2, 31, 40, 6)

for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 25
ws2.column_dimensions["A"].width = 14
ws2.freeze_panes = "A5"

# ============================================================
# ABA 3: PLANEJAMENTO SEMANAL
# ============================================================
ws3 = wb.create_sheet("Planejamento Semanal")
ws3.sheet_properties.tabColor = AZUL_CLARO

ws3.merge_cells("A1:F1")
ws3["A1"] = "📆 PLANEJAMENTO SEMANAL"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

plan_headers = ["Semana de", "Prioridade 1 (Operação)", "Prioridade 2 (Demandas)", "Prioridade 3 (Projetos)", "Resultado da Semana", "Auto-avaliação (1-5)"]
for i, h in enumerate(plan_headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 6)

ws3.cell(row=4, column=1, value="25/06/2026")
ws3.cell(row=4, column=2, value="Mapear processos e ingestões ativas")
ws3.cell(row=4, column=3, value="Verificar SDIBRE-1149 com Victor")
ws3.cell(row=4, column=4, value="Levantamento estado atual DataCore")

style_data_rows(ws3, 4, 30, 6)

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 40
ws3.column_dimensions["C"].width = 40
ws3.column_dimensions["D"].width = 40
ws3.column_dimensions["E"].width = 40
ws3.column_dimensions["F"].width = 22
ws3.freeze_panes = "A4"

# ============================================================
# ABA 4: MAPA DE RISCOS
# ============================================================
ws4 = wb.create_sheet("Mapa de Riscos")
ws4.sheet_properties.tabColor = VERMELHO

ws4.merge_cells("A1:G1")
ws4["A1"] = "⚠️ MAPA DE RISCOS E PROBLEMAS"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")

risk_headers = ["Data", "Risco / Problema", "Impacto", "Probabilidade", "Ação de Mitigação", "Status", "Responsável"]
for i, h in enumerate(risk_headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, 7)

# Pre-filled risks
risks = [
    ["25/06/2026", "Falta de envio de dados por informante", "Alto", "Média", "Acionar Flavio antes de tomar decisão", "Ativo", "Murilo + Flavio"],
    ["25/06/2026", "Atraso em demanda TI-IBRE (SDIBRE-1149)", "Médio", "Média", "Acompanhar semanalmente com Victor", "Ativo", "Murilo"],
    ["25/06/2026", "Perda de contexto na transição", "Alto", "Alta", "Documentar tudo + conversar com Brendon/Olivio nas primeiras semanas", "Ativo", "Murilo"],
    ["25/06/2026", "Sobrecarga operacional impedindo projetos", "Médio", "Alta", "Manter blocos protegidos para projetos (ter/qui tarde)", "Ativo", "Murilo"],
]

for row_idx, risk in enumerate(risks, 4):
    for col_idx, val in enumerate(risk, 1):
        ws4.cell(row=row_idx, column=col_idx, value=val)
    # Color code impact
    impact_cell = ws4.cell(row=row_idx, column=3)
    if impact_cell.value == "Alto":
        impact_cell.fill = red_fill
    elif impact_cell.value == "Médio":
        impact_cell.fill = yellow_fill

style_data_rows(ws4, 4, 20, 7)
# Re-apply risk colors after styling
for row_idx in range(4, 8):
    impact_cell = ws4.cell(row=row_idx, column=3)
    if impact_cell.value == "Alto":
        impact_cell.fill = red_fill
        impact_cell.font = Font(name="Calibri", bold=True, size=10, color=VERMELHO)
    elif impact_cell.value == "Médio":
        impact_cell.fill = yellow_fill
        impact_cell.font = Font(name="Calibri", bold=True, size=10, color=AMARELO)

ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 45
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 50
ws4.column_dimensions["F"].width = 12
ws4.column_dimensions["G"].width = 22
ws4.freeze_panes = "A4"

# ============================================================
# ABA 5: REGISTRO DE DECISÕES
# ============================================================
ws5 = wb.create_sheet("Registro de Decisões")
ws5.sheet_properties.tabColor = AZUL_MEDIO

ws5.merge_cells("A1:F1")
ws5["A1"] = "📝 REGISTRO DE DECISÕES"
ws5["A1"].font = title_font
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")

dec_headers = ["Data", "Decisão", "Contexto", "Alternativas Consideradas", "Quem foi Consultado", "Impacto"]
for i, h in enumerate(dec_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 6)
style_data_rows(ws5, 4, 25, 6)

ws5.column_dimensions["A"].width = 14
ws5.column_dimensions["B"].width = 40
ws5.column_dimensions["C"].width = 40
ws5.column_dimensions["D"].width = 35
ws5.column_dimensions["E"].width = 25
ws5.column_dimensions["F"].width = 30
ws5.freeze_panes = "A4"

# ============================================================
# ABA 6: PROJETOS
# ============================================================
ws6 = wb.create_sheet("Projetos")
ws6.sheet_properties.tabColor = "8E44AD"

ws6.merge_cells("A1:H1")
ws6["A1"] = "💻 ACOMPANHAMENTO DE PROJETOS"
ws6["A1"].font = title_font
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")

proj_headers = ["Projeto", "Status Geral", "% Conclusão", "Próximo Marco", "Prazo", "Bloqueios", "Horas Dedicadas (Semana)", "Notas"]
for i, h in enumerate(proj_headers, 1):
    ws6.cell(row=3, column=i, value=h)
style_header_row(ws6, 3, 8)

projects = [
    ["DataCore", "Em levantamento", "0%", "Entender estado atual do código e deploy", "09/07/2026", "—", "0", "Continuidade do desenvolvimento sob minha responsabilidade"],
    ["[Projeto Secundário 1]", "", "", "", "", "", "", ""],
    ["[Projeto Secundário 2]", "", "", "", "", "", "", ""],
    ["[Projeto Secundário 3]", "", "", "", "", "", "", ""],
]

for row_idx, proj in enumerate(projects, 4):
    for col_idx, val in enumerate(proj, 1):
        ws6.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws6, 4, 15, 8)

ws6.column_dimensions["A"].width = 25
ws6.column_dimensions["B"].width = 18
ws6.column_dimensions["C"].width = 14
ws6.column_dimensions["D"].width = 40
ws6.column_dimensions["E"].width = 14
ws6.column_dimensions["F"].width = 25
ws6.column_dimensions["G"].width = 24
ws6.column_dimensions["H"].width = 40
ws6.freeze_panes = "A4"

# ============================================================
# ABA 7: APRENDIZADOS
# ============================================================
ws7 = wb.create_sheet("Aprendizados")
ws7.sheet_properties.tabColor = VERDE

ws7.merge_cells("A1:E1")
ws7["A1"] = "🧠 COISAS QUE APRENDI / REFERÊNCIAS"
ws7["A1"].font = title_font
ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")

learn_headers = ["Data", "Assunto", "Detalhes", "Fonte / Quem ensinou", "Útil para"]
for i, h in enumerate(learn_headers, 1):
    ws7.cell(row=3, column=i, value=h)
style_header_row(ws7, 3, 5)

learnings = [
    ["25/06/2026", "Padrão de nomenclatura Snowflake", "Verificar exemplos existentes antes de criar novos objetos", "Verificar com equipe", "Criação de tabelas, schemas, procedures"],
    ["", "Processo de ingestão AWS → Snowflake", "Solicitar via TI-IBRE", "Brendon / Olivio", "Todas as ingestões"],
    ["", "Fluxo de carga Banco de Preços", "Cargas são responsabilidade do Flavio", "Flavio", "Entrega de dados ao BP"],
    ["", "Provisão de ambiente (SFTP/AWS)", "Levantar: tipo dado, sensível, formato, volumetria", "Documento de transição", "Novas parcerias"],
]

for row_idx, item in enumerate(learnings, 4):
    for col_idx, val in enumerate(item, 1):
        ws7.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws7, 4, 30, 5)

ws7.column_dimensions["A"].width = 14
ws7.column_dimensions["B"].width = 35
ws7.column_dimensions["C"].width = 50
ws7.column_dimensions["D"].width = 30
ws7.column_dimensions["E"].width = 40
ws7.freeze_panes = "A4"

# ============================================================
# ABA 8: CONTATOS
# ============================================================
ws8 = wb.create_sheet("Contatos")
ws8.sheet_properties.tabColor = AZUL_ESCURO

ws8.merge_cells("A1:F1")
ws8["A1"] = "📞 CONTATOS E REFERÊNCIAS RÁPIDAS"
ws8["A1"].font = title_font
ws8["A1"].alignment = Alignment(horizontal="center", vertical="center")

contact_headers = ["Pessoa", "Cargo / Papel", "Área", "Quando Acionar", "Canal Preferido", "Observações"]
for i, h in enumerate(contact_headers, 1):
    ws8.cell(row=3, column=i, value=h)
style_header_row(ws8, 3, 6)

contacts = [
    ["Flavio", "Gestão Banco de Preços", "SPDO", "Dúvidas operacionais / cargas / falta de dados de informante", "", "Reporte direto na ausência do responsável"],
    ["Arnaldo", "Reporte alternativo", "SPDO", "Na ausência do Flavio", "", ""],
    ["Victor", "TI-IBRE", "TI", "Demandas técnicas / acompanhamento SDIBRE-1149", "", ""],
    ["Rafaella Bezerra", "TI-IBRE", "TI", "Abertura de novas demandas", "", ""],
    ["Olivio", "Equipe (apoio ao Flavio)", "SPDO", "Governança / Cadastro / Banco de Preços", "", "Transição gradual para apoiar Flavio"],
    ["Brendon", "Equipe operacional", "SPDO", "Atividades operacionais", "", ""],
]

for row_idx, contact in enumerate(contacts, 4):
    for col_idx, val in enumerate(contact, 1):
        ws8.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws8, 4, 15, 6)

ws8.column_dimensions["A"].width = 20
ws8.column_dimensions["B"].width = 28
ws8.column_dimensions["C"].width = 10
ws8.column_dimensions["D"].width = 50
ws8.column_dimensions["E"].width = 20
ws8.column_dimensions["F"].width = 45
ws8.freeze_panes = "A4"

# ============================================================
# ABA 9: MONITORAMENTO DE DADOS
# ============================================================
ws9 = wb.create_sheet("Monitoramento de Dados")
ws9.sheet_properties.tabColor = AMARELO

ws9.merge_cells("A1:K1")
ws9["A1"] = "📊 PAINEL DE MONITORAMENTO DE DADOS"
ws9["A1"].font = title_font
ws9["A1"].alignment = Alignment(horizontal="center", vertical="center")

mon_headers = ["Fonte / Parceiro", "Tipo", "Categoria", "Formato", "Periodicidade", "Última Recepção", "Próxima Esperada", "Status", "Canal", "Contato Responsável", "Observações"]
for i, h in enumerate(mon_headers, 1):
    ws9.cell(row=3, column=i, value=h)
style_header_row(ws9, 3, 11)

# Section: Parcerias
ws9.cell(row=4, column=1, value="— PARCERIAS —")
ws9.cell(row=4, column=1).font = bold_font
sources = [
    ["Neogrid/Horus", "Parceria", "Dados de Parceria", "CSV", "Mensal", "", "", "Verificar", "SFTP", "", ""],
    ["Pastorinho", "Parceria", "Dados de Parceria", "CSV", "", "", "", "Verificar", "", "", ""],
    ["Starian", "Parceria", "Dados de Parceria", "CSV", "", "", "", "Verificar", "", "", ""],
    ["Divino Imóveis", "Parceria", "Dados Imobiliários", "", "", "", "", "Em andamento", "", "", ""],
    ["Moraes Administração", "Parceria", "Dados Imobiliários", "", "", "", "", "Em andamento", "", "", "Prazo: 30/06"],
]
for row_idx, src in enumerate(sources, 5):
    for col_idx, val in enumerate(src, 1):
        ws9.cell(row=row_idx, column=col_idx, value=val)

# Section: Tabelas de Coleta / Informantes
ws9.cell(row=11, column=1, value="— TABELAS DE INFORMANTES —")
ws9.cell(row=11, column=1).font = bold_font
informantes = [
    ["2373 - Wirex Cable", "Informante", "Tabela de Coleta", "", "B0110, M10, S0310, T0310", "", "", "", "", "", ""],
    ["68229 - TDM", "Informante", "Tabela de Coleta", "", "T0310", "", "", "", "", "", "Mar, Jun, Set, Dez (D1)"],
    ["103851 - Saint-Gobain/Tekbond", "Informante", "Tabela de Coleta", "", "T0110", "", "", "", "", "", "Jan, Abr, Jul, Out (D1)"],
    ["2533 - Saint-Gobain", "Informante", "Tabela de Coleta", "", "", "", "", "", "", "", "Gerar cargas + Importar no BP"],
]
for row_idx, inf in enumerate(informantes, 12):
    for col_idx, val in enumerate(inf, 1):
        ws9.cell(row=row_idx, column=col_idx, value=val)

# Section: Dados Externos
ws9.cell(row=17, column=1, value="— DADOS EXTERNOS —")
ws9.cell(row=17, column=1).font = bold_font
externos = [
    ["Base de CEP", "Externo", "Dados Externos", "", "", "", "", "", "", "Fonte pública", ""],
    ["Licitações", "Externo", "Dados Externos", "", "", "", "", "", "", "Fonte governamental", ""],
    ["COFOG", "Externo", "Dados Externos", "", "", "", "", "", "", "Fonte governamental", ""],
    ["RAIS", "Externo", "Dados Externos", "CSV", "Anual", "", "", "", "", "Min. Trabalho", ""],
    ["CAGED", "Externo", "Dados Externos", "CSV", "Mensal", "", "", "", "", "Min. Trabalho", ""],
]
for row_idx, ext in enumerate(externos, 18):
    for col_idx, val in enumerate(ext, 1):
        ws9.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws9, 4, 30, 11)

ws9.column_dimensions["A"].width = 30
ws9.column_dimensions["B"].width = 14
ws9.column_dimensions["C"].width = 20
ws9.column_dimensions["D"].width = 10
ws9.column_dimensions["E"].width = 28
ws9.column_dimensions["F"].width = 16
ws9.column_dimensions["G"].width = 16
ws9.column_dimensions["H"].width = 14
ws9.column_dimensions["I"].width = 12
ws9.column_dimensions["J"].width = 22
ws9.column_dimensions["K"].width = 30
ws9.freeze_panes = "A4"

# ============================================================
# ABA 10: BACKLOG DE DEMANDAS
# ============================================================
ws10 = wb.create_sheet("Backlog de Demandas")
ws10.sheet_properties.tabColor = "E74C3C"

ws10.merge_cells("A1:L1")
ws10["A1"] = "📋 BACKLOG DE DEMANDAS"
ws10["A1"].font = title_font
ws10["A1"].alignment = Alignment(horizontal="center", vertical="center")

dem_headers = ["ID", "Data Entrada", "Solicitante", "Categoria", "Descrição", "Tipo Atividade", "Prioridade", "Status", "Data Início", "Prazo", "Data Conclusão", "Observações"]
for i, h in enumerate(dem_headers, 1):
    ws10.cell(row=3, column=i, value=h)
style_header_row(ws10, 3, 12)

ws10.cell(row=4, column=1, value="DEM-001")
ws10.cell(row=4, column=2, value="25/06/2026")
ws10.cell(row=4, column=3, value="—")
ws10.cell(row=4, column=4, value="TI-IBRE")
ws10.cell(row=4, column=5, value="Verificar andamento SDIBRE-1149")
ws10.cell(row=4, column=6, value="Acompanhamento")
ws10.cell(row=4, column=7, value="Alta")
ws10.cell(row=4, column=8, value="Pendente")
ws10.cell(row=4, column=12, value="Falar com Victor")

ws10.cell(row=5, column=1, value="DEM-002")
ws10.cell(row=5, column=2, value="25/06/2026")
ws10.cell(row=5, column=4, value="Cloud Platform")
ws10.cell(row=5, column=5, value="Reestruturação Snowflake Schemas DB_IBRE")
ws10.cell(row=5, column=6, value="Infraestrutura")
ws10.cell(row=5, column=7, value="Média")
ws10.cell(row=5, column=8, value="Em andamento")
ws10.cell(row=5, column=12, value="Abrir demanda de correção com TI")

style_data_rows(ws10, 4, 30, 12)

ws10.column_dimensions["A"].width = 10
ws10.column_dimensions["B"].width = 14
ws10.column_dimensions["C"].width = 18
ws10.column_dimensions["D"].width = 16
ws10.column_dimensions["E"].width = 45
ws10.column_dimensions["F"].width = 18
ws10.column_dimensions["G"].width = 12
ws10.column_dimensions["H"].width = 18
ws10.column_dimensions["I"].width = 14
ws10.column_dimensions["J"].width = 14
ws10.column_dimensions["K"].width = 16
ws10.column_dimensions["L"].width = 35
ws10.freeze_panes = "A4"

# ============================================================
# SALVAR
# ============================================================
output_path = r"c:\Users\murilo.flores\Desktop\projetos\streamlit\projeto_fgv_repositorio_apps\docs\planilhas\controle_pessoal_murilo.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Arquivo salvo em: {output_path}")
